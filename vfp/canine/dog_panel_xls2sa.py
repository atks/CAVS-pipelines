#!/usr/bin/env python3

# The MIT License
# Copyright (c) 2021 Adrian Tan <adrian_tan@nparks.gov.sg>
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the 'Software'), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import os
import argparse
import textwrap
import re
import openpyxl


def main():
    cwd = os.getcwd()

    parser = argparse.ArgumentParser(
        description="Convert excel file to sample file",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent(
            """\
           usage: dog_panel_xls2sa.py <xlsx_file>
           """
        ),
    )
    parser.add_argument(
        "xlsx_file", type=str, help="Excel work sheet containing sample data"
    )
    parser.add_argument(
        "-s", "--sample_file", type=str, default="dogs.sa", help="Output sample file"
    )
    args = parser.parse_args()

    for arg in vars(args):
        print("\t{0:<20} :   {1:<10}".format(arg, getattr(args, arg) or ""))

    workbook = openpyxl.load_workbook(args.xlsx_file)
    dogs = []

    for sheet in workbook.worksheets:
        # print(sheet)
        for i in range(3, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value != None:
                id = sheet.cell(row=i, column=7).value
                type = sheet.cell(row=i, column=3).value
                breed = sheet.cell(row=i, column=4).value
                sex = sheet.cell(row=i, column=5).value
                colour = sheet.cell(row=i, column=6).value
                microchip = sheet.cell(row=i, column=2).value
                dogs.append(Dog(id, type, breed, sex, colour, microchip))

    # write to file
    file = open(args.sample_file, "w")
    file.write("#id\ttype\tsex\tbreed\tcolour\tmicrochip\n")
    for dog in dogs:
        file.write(dog.to_string() + "\n")

    file.close()


class Dog(object):
    def __init__(self, id, type, breed, sex, colour, microchip):
        self.id = self.normalise_id(id)
        self.type = type
        self.sex = sex.strip().title()
        self.breed = breed.strip().title().replace("( ", "(").replace(" )", ")")
        self.colour = colour.strip().title()
        self.microchip = microchip

    def normalise_id(self, id):
        m = re.match("([a-zA-Z]+)\s*(\d+)", id)
        #        print(id + ' -> ', end='')
        #        print(m.group(1) + '-', end='')
        #        print(m.group(2).lstrip('0').rjust(4,'0'))
        return m.group(1) + m.group(2).lstrip("0").rjust(4, "0")

    def print_row(self):
        print(
            f"{self.id}\t{self.type}\t{self.sex}\t{self.breed}\t"
            + f"{self.colour}\t{self.microchip}"
        )

    def to_string(self):
        return f"{self.id}\t{self.type}\t{self.sex}\t{self.breed}\t{self.colour}\t{self.microchip}"

    def print(self):
        print(f"id: {self.id}")
        print(f"type: {self.type}")
        print(f"sex: {self.sex}")
        print(f"breed: {self.breed}")
        print(f"colour: {self.colour}")
        print(f"microchip: {self.microchip}")


main()

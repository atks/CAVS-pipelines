#!/usr/bin/env python3

# The MIT License
# Copyright (c) 2021 Adrian Tan <adrian_tan@nparks.gov.sg>
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the 'Software'), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import os
import argparse
import textwrap
import re
import openpyxl


def main():
    cwd = os.getcwd()

    parser = argparse.ArgumentParser(
        description="Split up the directory FP into 2 based on sample dates",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent(
            """\
           usage: split_FP.py <xlsx_file>
           """
        ),
    )
    args = parser.parse_args()

    for arg in vars(args):
        print("\t{0:<20} :   {1:<10}".format(arg, getattr(args, arg) or ""))

    data_dir = "/data/vfp/canine_panel/20211118/data"
    info_dir = "/data/vfp/canine_panel/20211118/info"

    xlsx_file = f"{info_dir}/20211115_canine_panel.xlsx"
    workbook = openpyxl.load_workbook(xlsx_file)

    os.makedirs(
        "/data/vfp/canine_panel/20211118/data/Family Pet House (28 Jan)", exist_ok=True
    )
    os.makedirs(
        "/data/vfp/canine_panel/20211118/data/Family Pet House (29 Jan)", exist_ok=True
    )

    dog28 = dict()
    sheet = workbook["Family Pet House (28 Jan)"]
    for i in range(3, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value != None:
            id = sheet.cell(row=i, column=7).value
            dog_type = sheet.cell(row=i, column=3).value
            breed = sheet.cell(row=i, column=4).value
            sex = sheet.cell(row=i, column=5).value
            colour = sheet.cell(row=i, column=6).value
            microchip = sheet.cell(row=i, column=2).value
            id = normalise_id(id)
            #            print(id)
            #            dog28[id] = Dog(id, dog_type, breed, sex, colour, microchip)
            dog28[id] = 0

    dog29 = dict()
    sheet = workbook["Family Pet House (29 Jan)"]
    for i in range(3, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value != None:
            id = sheet.cell(row=i, column=7).value
            dog_type = sheet.cell(row=i, column=3).value
            breed = sheet.cell(row=i, column=4).value
            sex = sheet.cell(row=i, column=5).value
            colour = sheet.cell(row=i, column=6).value
            microchip = sheet.cell(row=i, column=2).value
            id = normalise_id(id)
            #            print(id)
            #            dog29[id] = Dog(id, dog_type, breed, sex, colour, microchip)
            dog29[id] = 0

    fp_dir = f"{data_dir}/FP"
    for subdir, dirs, files in os.walk(fp_dir):
        for file in files:
            if re.search("\.fsa\.csv$", file) != None:
                # check if corresponding fsa file exists
                base, ext = os.path.splitext(os.path.basename(file))
                if not os.path.isfile(f"{subdir}/{base}"):
                    exit("Does not exist")
                fsa_file = f"{subdir}/{base}"

                tokens = file.split("_")
                id = tokens[1]
                m = re.match("([a-zA-Z]+)\s*(\d+)(.*)", id)
                postamble = ""

                if m is not None:
                    id_norm = m.group(1).upper() + m.group(2).lstrip("0").rjust(4, "0")

                    if len(m.group(3)) != 0:
                        postamble = "_" + m.group(3).strip(" ()")
                #                        print(f"'{tokens[1]}'")
                #                        print(f"'{m.group(3)}'")
                #                        print(f"'{postamble}'")

                else:
                    # print(file)
                    if id == "FP MC4532":
                        id_norm = "FP0181"
                    elif id == "FP MC4534":
                        id_norm = "FP0182"
                    else:
                        print(f"\tNOT EXPECTED: {id}")

                if id_norm in dog28:
                    dog28[id_norm] += 1

                    print(
                        f"cp '{fsa_file}' '{data_dir}/Family Pet House (28 Jan)/{tokens[0]}_{id_norm}{postamble}_{tokens[2]}_{tokens[3]}.fsa'"
                    )
                    print(
                        f"cp '{fp_dir}/{file}' '{data_dir}/Family Pet House (28 Jan)/{tokens[0]}_{id_norm}{postamble}_{tokens[2]}_{tokens[3]}.fsa.csv'"
                    )

                if id_norm in dog29:
                    dog29[id_norm] += 1
                    print(
                        f"cp '{fsa_file}' '{data_dir}/Family Pet House (29 Jan)/{tokens[0]}_{id_norm}{postamble}_{tokens[2]}_{tokens[3]}.fsa'"
                    )
                    print(
                        f"cp '{fp_dir}/{file}' '{data_dir}/Family Pet House (29 Jan)/{tokens[0]}_{id_norm}{postamble}_{tokens[2]}_{tokens[3]}.fsa.csv'"
                    )


# NOTE
# FP0018 and FP0091 missing
# duplicates of FP0045 and FP0114


#    for id in dog28:
#        if dog28[id]!=1:
#            print(f"WEIRD: {id} {dog28[id]}")
#    for id in dog29:
#        if dog29[id]!=1:
#            print(f"WEIRD: {id} {dog29[id]}")


def normalise_id(id):
    m = re.match("([a-zA-Z]+)\s*(\d+)", id)
    return m.group(1).upper() + m.group(2).lstrip("0").rjust(4, "0")


class Dog(object):
    def __init__(self, id, type, breed, sex, colour, microchip):
        self.id = self.normalise_id(id)
        self.type = type
        self.sex = sex.strip().title()
        self.breed = breed.strip().title().replace("( ", "(").replace(" )", ")")
        self.colour = colour.strip().title()
        self.microchip = microchip

    def normalise_id(self, id):
        m = re.match("([a-zA-Z]+)\s*(\d+)", id)
        #        print(id + ' -> ', end='')
        #        print(m.group(1) + '-', end='')
        #        print(m.group(2).lstrip('0').rjust(4,'0'))
        return m.group(1).upper() + m.group(2).lstrip("0").rjust(4, "0")

    def print_row(self):
        print(
            f"{self.id}\t{self.type}\t{self.sex}\t{self.breed}\t"
            + f"{self.colour}\t{self.microchip}"
        )

    def to_string(self):
        return f"{self.id}\t{self.type}\t{self.sex}\t{self.breed}\t{self.colour}\t{self.microchip}"

    def print(self):
        print(f"id: {self.id}")
        print(f"type: {self.type}")
        print(f"sex: {self.sex}")
        print(f"breed: {self.breed}")
        print(f"colour: {self.colour}")
        print(f"microchip: {self.microchip}")

    def println(self):
        print(
            f"{self.id} {self.type} {self.sex} {self.breed} {self.colour} {self.microchip}"
        )


main()

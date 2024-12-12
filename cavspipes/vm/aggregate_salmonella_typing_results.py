#!/usr/bin/env python3

# The MIT License
# Copyright (c) 2024 Adrian Tan <adrian_tan@nparks.gov.sg>
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the 'Software'), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import os
import click
from shutil import copy2, which
import openpyxl
from openpyxl.workbook.views import BookView
from openpyxl.worksheet.table import Table, TableStyleInfo

@click.command()
@click.option(
    "-i",
    "--input_dir",
    show_default=True,
    required=True,
    help="Input Directory",
)
@click.option("-s", "--sample_file", required=True, help="sample file")
@click.option(
    "-o",
    "--output_xlsx",
    show_default=True,
    required=True,
    help="output xlsx file",
)
def main(input_dir, sample_file, output_xlsx):
    """
    Aggregate results from:
        a. seroseq2 v1.3.1
        b. SISTR 1.1.2
        c. mlst 2.23.0

    e.g. aggregate_salmonella_typing_pipeline.py -i salomnella_pt -s salmonella_pt.sa -o results.xlsx
    """
    if not os.path.isabs(input_dir):
        input_dir = f"{os.getcwd()}/{input_dir}"

    print("\t{0:<20} :   {1:<10}".format("make_file", input_dir))
    print("\t{0:<20} :   {1:<10}".format("sample_file", sample_file))
    print("\t{0:<20} :   {1:<10}".format("output_xlsx", output_xlsx))

    # read sample file
    samples = []
    with open(sample_file, "r") as file:
        index = 0
        for line in file:
            if not line.startswith("#"):
                index += 1
                sample_id, fastq1, fastq2, contigs_fasta = line.rstrip().split("\t")
                samples.append(
                    Sample(index, sample_id, fastq1, fastq2, contigs_fasta)
                )

    no_samples = len(samples)

    # write out xlsx
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    final_ws = wb.create_sheet("final")
    sistr_ws = wb.create_sheet("sistr")
    mlst_ws = wb.create_sheet("mlst")
    cgmlst_ws = wb.create_sheet("cgmlst")
    seqsero2_ws = wb.create_sheet("seqsero2")

    #set opening window size
    view = [BookView(xWindow=8000, yWindow=4000, windowWidth=25000, windowHeight=20000)]
    wb.views = view

    #table settings
    style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)

    tab = Table(displayName="final", ref=f"A1:G{no_samples+1}")
    tab.tableStyleInfo = style
    final_ws.add_table(tab)
    final_ws.sheet_view.zoomScale = 200

    tab = Table(displayName="sistr", ref=f"A1:H{no_samples+1}")
    tab.tableStyleInfo = style
    sistr_ws.add_table(tab)
    sistr_ws.sheet_view.zoomScale = 200

    tab = Table(displayName="mlst", ref=f"A1:J{no_samples+1}")
    tab.tableStyleInfo = style
    mlst_ws.add_table(tab)
    mlst_ws.sheet_view.zoomScale = 200

    tab = Table(displayName="cgmlst", ref=f"A1:H{no_samples+1}")
    tab.tableStyleInfo = style
    cgmlst_ws.add_table(tab)
    cgmlst_ws.sheet_view.zoomScale = 200

    tab = Table(displayName="seqsero2", ref=f"A1:I{no_samples+1}")
    tab.tableStyleInfo = style
    seqsero2_ws.add_table(tab)
    seqsero2_ws.sheet_view.zoomScale = 200

    #setup headers
    final_ws.cell(1, column=1).value = "sample"
    final_ws.cell(1, column=2).value = "sequence type (mlst)"
    final_ws.cell(1, column=3).value = "antigen profile (sistr)"
    final_ws.cell(1, column=4).value = "serogroup (sistr)"
    final_ws.cell(1, column=5).value = "serovar (sistr)"
    final_ws.cell(1, column=6).value = "species (seqsero2)"
    final_ws.cell(1, column=7).value = "note (seqsero2)"

    sistr_ws.cell(1, column=1).value = "sample"
    sistr_ws.cell(1, column=2).value = "o antigen"
    sistr_ws.cell(1, column=3).value = "h1 antigen"
    sistr_ws.cell(1, column=4).value = "h2 antigen"
    sistr_ws.cell(1, column=5).value = "serogroup"
    sistr_ws.cell(1, column=6).value = "serovar (consensus)"
    sistr_ws.cell(1, column=7).value = "serovar (antigen)"
    sistr_ws.cell(1, column=8).value = "serovar (cgmlst)"

    cgmlst_ws.cell(1, column=1).value = "sample"
    cgmlst_ws.cell(1, column=2).value = "sequence type"
    cgmlst_ws.cell(1, column=3).value = "distance"
    cgmlst_ws.cell(1, column=4).value = "found loci"
    cgmlst_ws.cell(1, column=5).value = "genome match"
    cgmlst_ws.cell(1, column=6).value = "matching alleles"
    cgmlst_ws.cell(1, column=7).value = "subspecies"
    cgmlst_ws.cell(1, column=8).value = "serovar"

    mlst_ws.cell(row=1, column=1).value = "sample"
    mlst_ws.cell(row=1, column=2).value = "mlst scheme"
    mlst_ws.cell(row=1, column=3).value = "sequence type"
    mlst_ws.cell(row=1, column=4).value = "aroC"
    mlst_ws.cell(row=1, column=5).value = "dnaN"
    mlst_ws.cell(row=1, column=6).value = "hemD"
    mlst_ws.cell(row=1, column=7).value = "hisD"
    mlst_ws.cell(row=1, column=8).value = "purE"
    mlst_ws.cell(row=1, column=9).value = "sucA"
    mlst_ws.cell(row=1, column=10).value = "thrA"

    seqsero2_ws.cell(1, column=1).value = "sample"
    seqsero2_ws.cell(1, column=2).value = "o antigen"
    seqsero2_ws.cell(1, column=3).value = "h1 antigen"
    seqsero2_ws.cell(1, column=4).value = "h2 antigen"
    seqsero2_ws.cell(1, column=5).value = "identity"
    seqsero2_ws.cell(1, column=6).value = "antigenic profile"
    seqsero2_ws.cell(1, column=7).value = "serotype"
    seqsero2_ws.cell(1, column=8).value = "serotype contamination"
    seqsero2_ws.cell(1, column=9).value = "note"

    # aggregate files
    for sample in samples:
        #sistr results
        with open(f"{input_dir}/{sample.id}/sistr/sistr.tab", "r") as file:
            sample_results = ""
            for line in file:
                sample_results = line.rstrip("\n")

            results = sample_results.split("\t")
            cgmlst_ST = results[0]
            cgmlst_distance = results[1]
            cgmlst_found_loci = results[2]
            cgmlst_genome_match = results[3]
            cgmlst_matching_alleles = results[4]
            cgmlst_subspecies = results[5]
            fasta_filepath = results[6]
            genome = results[7]
            h1 = results[8]
            h2 = results[9]
            o_antigen = results[10]
            serogroup = results[11]
            serovar = results[12]
            serovar_antigen = results[13]
            serovar_cgmlst = results[14]

            sistr_ws.cell(row=sample.idx+1, column=1).value = sample.id
            sistr_ws.cell(row=sample.idx+1, column=2).value = o_antigen
            sistr_ws.cell(row=sample.idx+1, column=3).value = h1
            sistr_ws.cell(row=sample.idx+1, column=4).value = h2
            sistr_ws.cell(row=sample.idx+1, column=5).value = serogroup
            sistr_ws.cell(row=sample.idx+1, column=6).value = serovar
            sistr_ws.cell(row=sample.idx+1, column=7).value = serovar_antigen
            sistr_ws.cell(row=sample.idx+1, column=8).value = serovar_cgmlst

            cgmlst_ws.cell(row=sample.idx+1, column=1).value = sample.id
            cgmlst_ws.cell(row=sample.idx+1, column=2).value = cgmlst_ST
            cgmlst_ws.cell(row=sample.idx+1, column=3).value = cgmlst_distance
            cgmlst_ws.cell(row=sample.idx+1, column=4).value = cgmlst_found_loci
            cgmlst_ws.cell(row=sample.idx+1, column=5).value = cgmlst_genome_match
            cgmlst_ws.cell(row=sample.idx+1, column=6).value = cgmlst_matching_alleles
            cgmlst_ws.cell(row=sample.idx+1, column=7).value = cgmlst_subspecies
            cgmlst_ws.cell(row=sample.idx+1, column=8).value = serovar_cgmlst

            final_ws.cell(row=sample.idx+1, column=1).value = sample.id
            final_ws.cell(row=sample.idx+1, column=3).value = f"{o_antigen}:{h1}:{h2}"
            final_ws.cell(row=sample.idx+1, column=4).value = serogroup
            final_ws.cell(row=sample.idx+1, column=5).value = serovar


        #mlst results
        with open(f"{input_dir}/{sample.id}/mlst/results.txt", "r") as file:
            sample_results = ""
            for line in file:
                sample_results = line.rstrip("\n")

            results = sample_results.split("\t")
            fasta_file = results[0]
            mlst_scheme = results[1]
            sequence_type = results[2]
            locus1_allele = results[3]
            locus2_allele = results[4]
            locus3_allele = results[5]
            locus4_allele = results[6]
            locus5_allele = results[7]
            locus6_allele = results[8]
            locus7_allele = results[9]

            mlst_ws.cell(row=sample.idx+1, column=1).value = sample.id
            mlst_ws.cell(row=sample.idx+1, column=2).value = mlst_scheme
            mlst_ws.cell(row=sample.idx+1, column=3).value = sequence_type
            mlst_ws.cell(row=sample.idx+1, column=4).value = locus1_allele
            mlst_ws.cell(row=sample.idx+1, column=5).value = locus2_allele
            mlst_ws.cell(row=sample.idx+1, column=6).value = locus3_allele
            mlst_ws.cell(row=sample.idx+1, column=7).value = locus4_allele
            mlst_ws.cell(row=sample.idx+1, column=8).value = locus5_allele
            mlst_ws.cell(row=sample.idx+1, column=9).value = locus6_allele
            mlst_ws.cell(row=sample.idx+1, column=10).value = locus7_allele

            final_ws.cell(row=sample.idx+1, column=2).value = sequence_type

        #seqsero2 results
        with open(f"{input_dir}/{sample.id}/seqsero2/SeqSero_result.tsv", "r") as file:
            sample_results = ""
            for line in file:
                sample_results = line.rstrip("\n")

            results = sample_results.split("\t")
            sample_name = results[0]
            output_dir = results[1]
            input_files = results[2]
            o_antigen = results[3]
            h1_antigen = results[4]
            h2_antigen = results[5]
            identity = results[6]
            antigenic_profile = results[7]
            serotype = results[8]
            serotype_contamination = results[9]
            note = results[10]

            seqsero2_ws.cell(row=sample.idx+1, column=1).value = sample.id
            seqsero2_ws.cell(row=sample.idx+1, column=2).value = o_antigen
            seqsero2_ws.cell(row=sample.idx+1, column=3).value = h1_antigen
            seqsero2_ws.cell(row=sample.idx+1, column=4).value = h2_antigen
            seqsero2_ws.cell(row=sample.idx+1, column=5).value = identity
            seqsero2_ws.cell(row=sample.idx+1, column=6).value = antigenic_profile
            seqsero2_ws.cell(row=sample.idx+1, column=7).value = serotype
            seqsero2_ws.cell(row=sample.idx+1, column=8).value = serotype_contamination
            seqsero2_ws.cell(row=sample.idx+1, column=9).value = note

            final_ws.cell(row=sample.idx+1, column=6).value = identity
            final_ws.cell(row=sample.idx+1, column=7).value = note

    for ws in wb.worksheets:
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

    wb.save(output_xlsx)
class Sample(object):
    def __init__(self, idx, id, fastq1, fastq2, contigs_fasta):
        self.idx = idx
        self.id = id
        self.fastq1 = fastq1
        self.fastq2 = fastq2
        self.contigs_fasta = contigs_fasta

    def print(self):
        print(f"index           : {self.idx}")
        print(f"id              : {self.id}")
        print(f"fastq1          : {self.fastq1}")
        print(f"fastq2          : {self.fastq2}")
        print(f"contigs fasta   : {self.contigs_fasta}")


if __name__ == "__main__":
    main() # type: ignore
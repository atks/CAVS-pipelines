#!/usr/bin/env python3

# The MIT License
# Copyright (c) 2024 Adrian Tan <adrian_tan@nparks.gov.sg>
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the 'Software'), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import os
import click
from shutil import copy2, which
import openpyxl
from openpyxl.workbook.views import BookView
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
import heapq

@click.command()
@click.option(
    "-i",
    "--input_dir",
    show_default=True,
    required=True,
    help="Input Directory",
)
@click.option("-s", "--sample_file", required=True, help="sample file")
@click.option(
    "-o",
    "--output_xlsx",
    show_default=True,
    required=True,
    help="output xlsx file",
)
def main(input_dir, sample_file, output_xlsx):
    """
    Aggregate results from:
        a. kraken2 results
        b. blast results

    e.g. aggregate_illu_results.py -i salomnella_pt -s salmonella_pt.sa -o summary.xlsx
    """
    input_dir = os.path.abspath(input_dir)
    sample_file = os.path.abspath(sample_file)
    output_xlsx = os.path.abspath(output_xlsx)

    #print("\t{0:<20} :   {1:<10}".format("make_file", input_dir))
    #print("\t{0:<20} :   {1:<10}".format("sample_file", sample_file))
    #print("\t{0:<20} :   {1:<10}".format("output_xlsx", output_xlsx))

    # read sample file
    samples = []
    with open(sample_file, "r") as file:
        index = 0
        for line in file:
            if not line.startswith("#"):
                index += 1
                sample_id, fastq1, fastq2 = line.rstrip().split("\t")
                if sample_id != "unclassified":
                    samples.append(
                        Sample(index, sample_id, fastq1, fastq2)
                    )

    no_samples = len(samples)

    # write out xlsx
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    final_ws = wb.create_sheet("final")
    blast_ws = wb.create_sheet("blast")
    kraken2_ws = wb.create_sheet("kraken2")

    #set opening window size
    view = [BookView(xWindow=8000, yWindow=4000, windowWidth=25000, windowHeight=20000)]
    wb.views = view

    #table settings
    style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)

    tab = Table(displayName="final", ref=f"A1:E{no_samples+1}")
    tab.tableStyleInfo = style
    final_ws.add_table(tab)
    final_ws.sheet_view.zoomScale = 200

    tab = Table(displayName="blast", ref=f"A1:K{no_samples+1}")
    tab.tableStyleInfo = style
    blast_ws.add_table(tab)
    blast_ws.sheet_view.zoomScale = 200

    tab = Table(displayName="kraken2", ref=f"A1:J{no_samples+1}")
    tab.tableStyleInfo = style
    kraken2_ws.add_table(tab)
    kraken2_ws.sheet_view.zoomScale = 200

    #setup headers
    final_ws.cell(1, column=1).value = "sample"
    final_ws.cell(1, column=2).value = "blast species"
    final_ws.cell(1, column=3).value = "blast species (%)"
    final_ws.cell(1, column=4).value = "kraken2 species"
    final_ws.cell(1, column=5).value = "kraken2 species (%)"

    blast_ws.cell(1, column=1).value = "sample"
    blast_ws.cell(1, column=2).value = "annotated contig count"
    blast_ws.cell(1, column=3).value = "annotation count (up to 10 unique sequences per contig)"
    blast_ws.cell(1, column=4).value = "species 1"
    blast_ws.cell(1, column=5).value = "species 1 (%)"
    blast_ws.cell(1, column=6).value = "species 2"
    blast_ws.cell(1, column=7).value = "species 2 (%)"
    blast_ws.cell(1, column=8).value = "species 3"
    blast_ws.cell(1, column=9).value = "species 3 (%)"
    blast_ws.cell(1, column=10).value = "other species (%)"
    blast_ws.cell(1, column=11).value = "other species"

    kraken2_ws.cell(1, column=1).value = "sample"
    kraken2_ws.cell(1, column=2).value = "species level read count"
    kraken2_ws.cell(1, column=3).value = "species 1"
    kraken2_ws.cell(1, column=4).value = "species 1 (%)"
    kraken2_ws.cell(1, column=5).value = "species 2"
    kraken2_ws.cell(1, column=6).value = "species 2 (%)"
    kraken2_ws.cell(1, column=7).value = "species 3"
    kraken2_ws.cell(1, column=8).value = "species 3 (%)"
    kraken2_ws.cell(1, column=9).value = "other species (%)"
    kraken2_ws.cell(1, column=10).value = "other species"

    # aggregate files
    for sample in samples:

        #blast results
        blast_txt_file = f"{input_dir}/analysis/{sample.idx}_{sample.id}/blast_result/{sample.padded_idx}_{sample.id}.txt"

        try:
            with open(blast_txt_file, "r") as file:
                #print(blast_txt_file)
                unique_contigs = {}
                species_count = {}

                for line in file:
                    results = line.rstrip("\n").split("\t")
                    contig = results[0]
                    #staxid = results[8]
                    ssciname = results[9]
                    #scomname = results[10]
                    #sskingdom = results[11]

                    unique_contigs[contig] = 1

                    if ssciname in species_count:
                        species_count[ssciname] += 1
                    else:
                        species_count[ssciname] = 1

                species_heap = []
                for name in species_count:
                    heapq.heappush(species_heap, Species(species_count[name], name))

                no_annotated_contigs = len(unique_contigs)

                # get top 3 species
                total_species_annotation_count = 0.0
                other_species = []
                for i, species in enumerate(species_heap):
                    total_species_annotation_count += species.no_reads
                    if i>2:
                        other_species.append(species.name)

                blast_ws.cell(row=sample.idx+1, column=1).value = f"{sample.idx}_{sample.id}"
                blast_ws.cell(row=sample.idx+1, column=2).value = no_annotated_contigs
                blast_ws.cell(row=sample.idx+1, column=3).value = total_species_annotation_count

                top_species = heapq.nlargest(3, species_heap)
                top_species_annotation_count = 0.0
                for i, species in enumerate(top_species):
                    blast_ws.cell(row=sample.idx+1, column=2*i+4).value = species.name
                    blast_ws.cell(row=sample.idx+1, column=2*i+5).value = f"{species.no_reads/total_species_annotation_count*100:.2f}"
                    top_species_annotation_count += species.no_reads
                    if i == 0:
                        final_ws.cell(row=sample.idx+1, column=2).value = species.name
                        final_ws.cell(row=sample.idx+1, column=3).value = f"{species.no_reads/total_species_annotation_count*100:.2f}"

                # get rest of species
                blast_ws.cell(row=sample.idx+1, column=10).value = f"{(total_species_annotation_count-top_species_annotation_count)/total_species_annotation_count*100:.2f}"
                blast_ws.cell(row=sample.idx+1, column=11).value = ":".join(other_species)
                final_ws.cell(row=sample.idx+1, column=1).value = f"{sample.idx}_{sample.id}"

        except FileNotFoundError as e:
            #print(f"File does not exist: {e.filename}")
            blast_ws.cell(row=sample.idx+1, column=1).value = f"{sample.idx}_{sample.id}"
            for i in range(2,12):
                blast_ws.cell(row=sample.idx+1, column=i).value = "n/a"
            final_ws.cell(row=sample.idx+1, column=1).value = f"{sample.idx}_{sample.id}"
            final_ws.cell(row=sample.idx+1, column=2).value = "n/a"
            final_ws.cell(row=sample.idx+1, column=3).value = "n/a"


        #kraken2 results
        kraken2_txt_file = f"{input_dir}/analysis/{sample.idx}_{sample.id}/kraken2_result/{sample.padded_idx}_{sample.id}.txt"
        try:
            with open(kraken2_txt_file, "r") as file:
                #print(kraken2_txt_file)
                species_heap = []
                for line in file:
                    percentage_reads, no_reads_clade, no_assigned_reads, tax_level, tax_id, nomenclature =  line.strip().split(maxsplit=5)
                    if tax_level == "S":
                        heapq.heappush(species_heap, Species(int(no_reads_clade), nomenclature))

                # get top 3 species
                total_species_read_count = 0.0
                other_species = []
                for i, species in enumerate(species_heap):
                    total_species_read_count += species.no_reads
                    if i>2:
                        other_species.append(species.name)

                kraken2_ws.cell(row=sample.idx+1, column=1).value = f"{sample.idx}_{sample.id}"
                kraken2_ws.cell(row=sample.idx+1, column=2).value = total_species_read_count

                top_species = heapq.nlargest(3, species_heap)
                top_species_read_count = 0.0
                for i, species in enumerate(top_species):
                    kraken2_ws.cell(row=sample.idx+1, column=2*i+3).value = species.name
                    kraken2_ws.cell(row=sample.idx+1, column=2*i+4).value = f"{species.no_reads/total_species_read_count*100:.2f}"
                    top_species_read_count += species.no_reads
                    if i == 0:
                        final_ws.cell(row=sample.idx+1, column=4).value = species.name
                        final_ws.cell(row=sample.idx+1, column=5).value = f"{species.no_reads/total_species_read_count*100:.2f}"

                # get rest of species
                kraken2_ws.cell(row=sample.idx+1, column=9).value = f"{(total_species_read_count-top_species_read_count)/total_species_read_count*100:.2f}"
                kraken2_ws.cell(row=sample.idx+1, column=10).value = ":".join(other_species)

        except OSError as e:
            print(f"Error: {kraken2_txt_file} : {e}")
            pass

    for ws in wb.worksheets:
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            print(f"{col} {value}")
            ws.column_dimensions[col].width = value
        print(f"==============")

    wb.save(output_xlsx)
    wb.close()


class Sample(object):
    def __init__(self, idx, id, fastq1, fastq2):
        self.idx = idx
        self.padded_idx = f"{idx:02}"
        self.id = id
        self.fastq1 = fastq1
        self.fastq2 = fastq2

    def print(self):
        print(f"index           : {self.idx}")
        print(f"padded index    : {self.padded_idx}")
        print(f"id              : {self.id}")
        print(f"fastq1          : {self.fastq1}")
        print(f"fastq2          : {self.fastq2}")

class Species(object):
    def __init__(self, no_reads, name):
        self.no_reads = no_reads
        self.name = name

    def __lt__(self, other):
        return self.no_reads < other.no_reads

if __name__ == "__main__":
    main() # type: ignore
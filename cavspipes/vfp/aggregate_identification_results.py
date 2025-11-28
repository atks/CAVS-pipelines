#!/usr/bin/env python3

# The MIT License
# Copyright (c) 2025 Adrian Tan <adrian_tan@nparks.gov.sg>
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the 'Software'), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import os
import click
from shutil import copy2, which
import openpyxl
from openpyxl.workbook.views import BookView
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
import heapq
import subprocess
import re
from pathlib import Path

@click.command()
@click.option(
    "-i",
    "--input_dir",
    show_default=True,
    required=True,
    help="Input Directory",
)
@click.option("-s", "--sample_file", required=True, help="sample file")
@click.option(
    "-o",
    "--output_xlsx",
    show_default=True,
    required=True,
    help="output xlsx file",
)
@click.option(
    "--suffix",
    show_default=True,
    default="",
    help="suffix for identification result directory",
)
def main(input_dir, sample_file, output_xlsx, suffix):
    """
    Aggregate results from identification of barcodes.

    e.g. aggregate_identification_results.py -i pore16 -s pore16_pt.sa -o summary.xlsx
    """
    input_dir = os.path.abspath(input_dir)
    sample_file = os.path.abspath(sample_file)
    output_xlsx = os.path.abspath(output_xlsx)
    output_dir = os.path.dirname(output_xlsx)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    #print("\t{0:<20} :   {1:<10}".format("make_file", input_dir))
    #print("\t{0:<20} :   {1:<10}".format("sample_file", sample_file))
    #print("\t{0:<20} :   {1:<10}".format("output_xlsx", output_xlsx))

    # read sample file
    samples = []
    with open(sample_file, "r") as file:
        index = 0
        for line in file:
            if not line.startswith("#"):
                index += 1
                sample_id, barcode, min_len, max_len = line.rstrip().split("\t")
                if sample_id != "unclassified":
                    samples.append(
                        Sample(index, sample_id, barcode, int(min_len), int(max_len))
                              )

    no_samples = len(samples)

    # write out xlsx
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    summary_ws = wb.create_sheet("summary")

    #set opening window size
    view = [BookView(xWindow=8000, yWindow=4000, windowWidth=25000, windowHeight=20000)]
    wb.views = view

    #table settings
    style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    summary_ws.sheet_view.zoomScale = 200

    # aggregate files
    for sample in samples:
        sample.collect_info(input_dir, suffix)
        #sample.print_contigs()

    # print to excel
    sample_row = 1
    for sample in samples:
        summary_ws.cell(sample_row, column=1).value = sample.id
        summary_ws.cell(sample_row, column=2).value = f"{sample.no_reads_in_length_range}/{sample.total_reads} ({sample.no_reads_in_length_range/sample.total_reads*100:.2f}%)"
        #headers
        summary_ws.cell(sample_row+1, column=1).value = "consensus contig"
        summary_ws.cell(sample_row+1, column=2).value = "#supporting reads"
        summary_ws.cell(sample_row+1, column=3).value = "Best match Species"
        summary_ws.cell(sample_row+1, column=4).value = "Best match accession"
        summary_ws.cell(sample_row+1, column=5).value = "Score"
        summary_ws.cell(sample_row+1, column=6).value = "Query Length"
        summary_ws.cell(sample_row+1, column=7).value = "Subject Length"
        summary_ws.cell(sample_row+1, column=8).value = "Overlap Length"
        summary_ws.cell(sample_row+1, column=9).value = "Query Cover"
        summary_ws.cell(sample_row+1, column=10).value = "Percentage Identity"
        summary_ws.cell(sample_row+1, column=11).value = "Rest of the hits"
        i = 2
        if len(sample.contigs) != 0:
            for contig in sample.contigs.values():
                summary_ws.cell(sample_row+i, column=1).value = contig.name
                summary_ws.cell(sample_row+i, column=2).value = contig.no_reads
                if len(contig.sorted_alignments) > 0:
                    alignment = heapq.heappop(contig.sorted_alignments)
                    summary_ws.cell(sample_row+i, column=3).value = alignment.sscinames
                    summary_ws.cell(sample_row+i, column=4).value = alignment.sacc
                    summary_ws.cell(sample_row+i, column=5).value = alignment.score
                    summary_ws.cell(sample_row+i, column=6).value = alignment.qlen
                    summary_ws.cell(sample_row+i, column=7).value = alignment.slen
                    summary_ws.cell(sample_row+i, column=8).value = alignment.length
                    summary_ws.cell(sample_row+i, column=9).value = f"{100.0*alignment.length/alignment.qlen:.2f}"
                    summary_ws.cell(sample_row+i, column=10).value = f"{alignment.pident:.2f}"
                    collated_hits = []
                    while len(contig.sorted_alignments) > 0:
                        alignment = heapq.heappop(contig.sorted_alignments)
                        collated_hits.append(f"{alignment.sscinames} ({alignment.score})")
                    summary_ws.cell(sample_row+i, column=11).value = f"{";".join(collated_hits)}"
                    i += 1
                else:
                    summary_ws.cell(sample_row+i, column=3).value = "No BLAST hits"
                    summary_ws.cell(sample_row+i, column=6).value = contig.length                   
                    i += 1    
        else:
            summary_ws.cell(sample_row+i, column=1).value = "No contigs assembled"
            i += 1

        tab = Table(displayName=f"{sample.id}", ref=f"A{sample_row+1}:K{sample_row+i-1}")
        tab.tableStyleInfo = style
        summary_ws.add_table(tab)

        #update rows
        sample_row += i + 1

    for ws in wb.worksheets:
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    wb.save(output_xlsx)
    wb.close()

class Sample(object):
    def __init__(self, idx, id, barcode, min_len, max_len):
        self.idx = idx
        self.padded_idx = f"{idx:02}"
        self.id = id
        self.barcode = barcode
        self.min_len = min_len
        self.max_len = max_len
        self.total_reads = 0
        self.no_reads_in_length_range = 0
        self.contigs = {}

    def collect_info(self, input_dir, suffix):
        nanoplot_txt_file = f"{input_dir}/analysis/{self.idx}_{self.id}/nanoplot_result/{self.padded_idx}_{self.id}.txt"
        identification_result_dir = f"{input_dir}/analysis/{self.idx}_{self.id}/identification_result{suffix}"
        ampliconsorter_csv_file = f"{identification_result_dir}/amplicon_sorter/results.csv"
        #consensus_fasta_file = f"{identification_result_dir}/amplicon_sorter/{self.idx}_{self.id}_consensussequences.fasta"
        consensus_fasta_file = f"{identification_result_dir}/amplicon_sorter/consensusfile.fasta"
        blast_txt_file = f"{identification_result_dir}/blast/{self.padded_idx}_{self.id}.txt"

        print(f"processing {nanoplot_txt_file}")
        print(f"processing {ampliconsorter_csv_file}")
        print(f"processing {consensus_fasta_file}")
        print(f"processing {blast_txt_file}")

        with open(nanoplot_txt_file, "r") as file:
            for line in file:
                if line.startswith("Number of reads:"):
                    self.total_reads = int(float(line.split()[-1].replace(',', '')))
                    break
        try:
            with open(ampliconsorter_csv_file, "r") as file:
                for line in file:
                    line = line.rstrip()
                    if line=="":
                        continue
                    contig_name, read_no = line.split(",")
                    if contig_name == "Total":
                        self.no_reads_in_length_range = int(read_no)
                    elif len(contig_name) > 0:
                        self.contigs[contig_name] = Contig(contig_name, 0, int(read_no))
        except FileNotFoundError as e:
            print(f"File does not exist: {e.filename}")

        path = Path(consensus_fasta_file)

        if not path.exists():
            return

        cmd = f"seqtk comp {consensus_fasta_file} | cut -f1,2"
        result = subprocess.run(cmd, shell=True, check=True, capture_output=True, text=True)      
        
        if result.stdout != "":
            lines = result.stdout.strip()
            for line in lines.split("\n"):
                #print(line)
                contig_name, length = line.removeprefix("consensus_").split("\t")
                contig_name = re.sub(r"\(\d+\)$", "", contig_name)
                if contig_name in self.contigs:
                    self.contigs[contig_name].length = int(length)
                else:
                    print(f"Contig {contig_name} not found in amplicon sorter results")

            with open(blast_txt_file, "r") as file:
                #populate alignments for each contig
                #if a contig is aligned more than once to the same sacc, only the best alignment is kept
                for line in file:
    #                 qacc sacc qlen slen score length pident stitle staxids sscinames scomnames sskingdoms
                    qacc, sacc, qlen, slen, score, length, pident, stitle, staxids, sscinames, scomnames, sskingdoms = line.rstrip("\n").split("\t")
                    contig_name = qacc.removeprefix("consensus_")
                    contig_name = re.sub(r"\(\d+\)$", "", contig_name)
                    #print(f"contig_name: {contig_name}")
                    if contig_name in self.contigs:
                        if sacc not in self.contigs[contig_name].alignments:
                            self.contigs[contig_name].alignments[sacc] = Alignment(qacc, sacc, qlen, slen, score, length, pident, stitle, staxids, sscinames, scomnames, sskingdoms)
                        else:
                            if int(score) > self.contigs[contig_name].alignments[sacc].score:
                                print(f"Updating alignment for {contig_name} with {sacc}")
                                self.contigs[contig_name].alignments[sacc].print()
                                self.contigs[contig_name].alignments[sacc] = Alignment(qacc, sacc, qlen, slen, score, length, pident, stitle, staxids, sscinames, scomnames, sskingdoms)
                                self.contigs[contig_name].alignments[sacc].print()
                    else:
                        print(f"Contig {contig_name} not found")

            #sort alignments for each contig
            for contig_name in self.contigs:
                #sort alignments by score
                self.contigs[contig_name].sorted_alignments = []
                for alignment in self.contigs[contig_name].alignments.values():
                    heapq.heappush(self.contigs[contig_name].sorted_alignments, alignment)
        

    def print_contigs(self):
        for name, contig in self.contigs.items():
            self.contigs[name].print()


    def print(self):
        print(f"index           : {self.idx}")
        print(f"padded index    : {self.padded_idx}")
        print(f"id              : {self.id}")
        print(f"barcode         : {self.barcode}")
        print(f"min_len         : {self.min_len}")
        print(f"max_len         : {self.max_len}")

class Contig(object):
    def __init__(self, name, length, reads):
        self.name = name
        self.length = length
        self.no_reads = reads
        self.alignments = {}
        self.sorted_alignments = []

    def print(self):
        print(f"contig name     : {self.name}")
        print(f"contig length   : {self.length}")
        print(f"contig reads    : {self.no_reads}")
        for alignment in self.alignments.values():
            alignment.print()

class Alignment(object):
    def __init__(self, qacc, sacc, qlen, slen, score, length, pident, stitle, staxids, sscinames, scomnames, sskingdoms):
        self.qacc = qacc
        self.sacc = sacc
        self.qlen = int(qlen)
        self.slen = int(slen)
        self.score = int(score)
        self.length = int(length)
        self.pident = float(pident)
        self.stitle = stitle
        self.staxids = staxids
        self.sscinames = sscinames
        self.scomnames = scomnames
        self.sskingdoms = sskingdoms

    def __lt__(self, other):
        return self.score > other.score

    def print(self):
        print(f"\t={self.sacc}=")
        print(f"\t\tslen       : {self.slen}")
        print(f"\t\tscore      : {self.score}")
        print(f"\t\tlength     : {self.length}")
        print(f"\t\tpident     : {self.pident}")
        print(f"\t\tstitle     : {self.stitle}")
        print(f"\t\tstaxids    : {self.staxids}")
        print(f"\t\tsscinames  : {self.sscinames}")
        print(f"\t\tscomnames  : {self.scomnames}")
        print(f"\t\tsskingdoms : {self.sskingdoms}")

class Species(object):
    def __init__(self, no_reads, name):
        self.no_reads = no_reads
        self.name = name

    def __lt__(self, other):
        return self.no_reads < other.no_reads

if __name__ == "__main__":
    main() # type: ignore
#!/usr/bin/env python3

# The MIT License
# Copyright (c) 2025 Adrian Tan <adrian_tan@nparks.gov.sg>
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the 'Software'), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import csv
import click
from shutil import copy2
from openpyxl import Workbook
from openpyxl import load_workbook

@click.command()
@click.option(
    "-t",
    "--txt_file",
    show_default=True,
    help="txt file",
)
@click.option(
    "-x",
    "--xlsx__file",
    show_default=True,
    help="xlsx file",
)
def main(txt_file, xlsx_file):
    """
    Convert excel file into text file

    e.g. formsg_csv_to_xlsx.py -t poe.txt -x poe_results.xlsx
    """
    print("\t{0:<20} :   {1:<10}".format("text file", txt_file))
    print("\t{0:<20} :   {1:<10}".format("excel file", xlsx_file))

    #open copy of the excel file
    wb = load_workbook(xlsx_file)
    ws = wb['For calculation']

    #read in text file
    i = 3
    with open(txt_file, mode='w') as file:
        for line in file:
            if not line.startswith("Response ID"):
                fields = line.strip().split("\t")
                #building and storey
                ws[f'B{i}'] = ""
                ws[f'C{i}'] = fields[3]
                #gender
                if fields[4] == "Male":
                    ws[f'F{i}'] = "1"
                else:
                    ws[f'F{i}'] = "2"
                #age group
                if fields[5] == ">61 years":
                    ws[f'I{i}'] = "6"
                elif fields[5] == "51-60 years":
                    ws[f'I{i}'] = "5"
                elif fields[5] == "41-50 years":
                    ws[f'I{i}'] = "4"
                elif fields[5] == "31-40 years":
                    ws[f'I{i}'] = "3"
                elif fields[5] == "21-30 years":
                    ws[f'I{i}'] = "2"
                #job category
                if fields[6].startswith("Others"):
                    ws[f'L{i}'] = "4"
                    ws[f'M{i}'] = fields[6].split(":")[1]
                else:
                    if fields[6] == "Managerial":
                        ws[f'L{i}'] = "1"
                    elif fields[6] == "Professional":
                        ws[f'L{i}'] = "2"
                    elif fields[6] == "Administrative":
                        ws[f'L{i}'] = "3"
                #job nature
                #hours spent
                if fields[8] == "< 10 hours":
                    ws[f'P{i}'] = "1"
                elif fields[8] == "11-15 hours":
                    ws[f'P{i}'] = "2"
                elif fields[8] == "16-20 hours":
                    ws[f'P{i}'] = "3"
                elif fields[8] == "21-25 hours":
                    ws[f'P{i}'] = "4"
                elif fields[8] == "26-30 hours":
                    ws[f'P{i}'] = "5"
                elif fields[8] == ">30 hours":
                    ws[f'P{i}'] = "6"
                #workplace type
                if fields[9].startswith("Enclosed"):
                    ws[f'S{i}'] = "1"
                else:
                    ws[f'S{i}'] = "2"

                #surrouding work environment
                if fields[10].startswith("Others"):
                    ws[f'V{i}'] = "6"
                    ws[f'W{i}'] = fields[10].split(":")[1]
                else:
                    surrounding_work_environment = fields[10].split(";")
                    if "Photocopier / Printer" in  surrounding_work_environment:
                        ws[f'V{i}'] = "1"
                    elif "Server Rack / Room" in  surrounding_work_environment:
                        ws[f'V{i}'] = "2"
                    elif "Pantry" in  surrounding_work_environment:
                        ws[f'V{i}'] = "3"
                    elif "Entrance" in  surrounding_work_environment:
                        ws[f'V{i}'] = "4"
                    elif "Not applicable" in  surrounding_work_environment:
                        ws[f'V{i}'] = "5"

                #Which of the following do you often use to attain thermal comfort?
                if fields[11].startswith("Others"):
                    ws[f'Z{i}'] = "5"
                    ws[f'AA{i}'] = fields[11].split(":")[1]
                else:
                    thermal_comfort = fields[11].split(";")
                    if "Fans" in  thermal_comfort:
                        ws[f'Z{i}'] = "1"
                    elif "Extra Clothes" in  thermal_comfort:
                        ws[f'Z{i}'] = "2"
                    elif "Less Clothes" in  thermal_comfort:
                        ws[f'Z{i}'] = "3"
                    elif "Not applicable" in  thermal_comfort:
                        ws[f'Z{i}'] = "4"

                #Do you experience an unpleasant odour?
                #Which of the following do you often use to attain thermal comfort?
                if fields[12] == "Regularly":
                    ws[f'AD{i}'] = "1"
                elif fields[12] == "Sometimes":
                    ws[f'AD{i}'] = "2"
                elif fields[12] == "Never":
                    ws[f'AD{i}'] = "3"

                #Do you have asthma?
                #Asthma - Are you on medication?
                if fields[13] == "Yes":
                    if fields[14] == "Yes":
                        ws[f'AG{i}'] = "Yes, on medication"
                    elif fields[14] == "No":
                        ws[f'AG{i}'] = "Yes, not on medication"
                elif fields[13] == "No":
                    ws[f'AG{i}'] = "No"

                # 16	Do you have an allergy?
                # 17	Allergy - Are you on medication?
                if fields[15] == "Yes":
                    if fields[16] == "Yes":
                        ws[f'AH{i}'] = "Yes, on medication"
                    elif fields[16] == "No":
                        ws[f'AH{i}'] = "Yes, not on medication"
                elif fields[15] == "No":
                    ws[f'AH{i}'] = "No"

                # 18	Do you have sinus issues?
                # 19	Sinus - Are you on medication?
                if fields[17] == "Yes":
                    if fields[18] == "Yes":
                        ws[f'AI{i}'] = "Yes, on medication"
                    elif fields[18] == "No":
                        ws[f'AI{i}'] = "Yes, not on medication"
                elif fields[17] == "No":
                    ws[f'AI{i}'] = "No"

                # 20	Do you have migraine issues?
                # 21	Migraine - Are you on medication?
                if fields[19] == "Yes":
                    if fields[20] == "Yes":
                        ws[f'AJ{i}'] = "Yes, on medication"
                    elif fields[20] == "No":
                        ws[f'AJ{i}'] = "Yes, not on medication"
                elif fields[19] == "No":
                    ws[f'AJ{i}'] = "No"

                # 22	Thermal Comfort
                if fields[21] == "Excellent":
                    ws[f'AM{i}'] = "1"
                elif fields[21] == "Good":
                    ws[f'AM{i}'] = "2"
                elif fields[21] == "Average":
                    ws[f'AM{i}'] = "3"
                elif fields[21] == "Poor":
                    ws[f'AM{i}'] = "4"
                elif fields[21] == "Very Poor":
                    ws[f'AM{i}'] = "5"

                # 23	Air Quality
                if fields[22] == "Excellent":
                    ws[f'AN{i}'] = "1"
                elif fields[22] == "Good":
                    ws[f'AN{i}'] = "2"
                elif fields[22] == "Average":
                    ws[f'AN{i}'] = "3"
                elif fields[22] == "Poor":
                    ws[f'AN{i}'] = "4"
                elif fields[22] == "Very Poor":
                    ws[f'AN{i}'] = "5"

                # 24	Lighting Level
                if fields[23] == "Excellent":
                    ws[f'AO{i}'] = "1"
                elif fields[23] == "Good":
                    ws[f'AO{i}'] = "2"
                elif fields[23] == "Average":
                    ws[f'AO{i}'] = "3"
                elif fields[23] == "Poor":
                    ws[f'AO{i}'] = "4"
                elif fields[23] == "Very Poor":
                    ws[f'AO{i}'] = "5"

                # 25	Daylight Level
                if fields[24] == "Excellent":
                    ws[f'AP{i}'] = "1"
                elif fields[24] == "Good":
                    ws[f'AP{i}'] = "2"
                elif fields[24] == "Average":
                    ws[f'AP{i}'] = "3"
                elif fields[24] == "Poor":
                    ws[f'AP{i}'] = "4"
                elif fields[24] == "Very Poor":
                    ws[f'AP{i}'] = "5"

                # 26	Window view to outside
                if fields[25] == "Excellent":
                    ws[f'AQ{i}'] = "1"
                elif fields[25] == "Good":
                    ws[f'AQ{i}'] = "2"
                elif fields[25] == "Average":
                    ws[f'AQ{i}'] = "3"
                elif fields[25] == "Poor":
                    ws[f'AQ{i}'] = "4"
                elif fields[25] == "Very Poor":
                    ws[f'AQ{i}'] = "5"

                # 27	Noise Level
                if fields[26] == "Excellent":
                    ws[f'AR{i}'] = "1"
                elif fields[26] == "Good":
                    ws[f'AR{i}'] = "2"
                elif fields[26] == "Average":
                    ws[f'AR{i}'] = "3"
                elif fields[26] == "Poor":
                    ws[f'AR{i}'] = "4"
                elif fields[26] == "Very Poor":
                    ws[f'AR{i}'] = "5"

                # 28	Overall cleanliness
                if fields[27] == "Excellent":
                    ws[f'AS{i}'] = "1"
                elif fields[27] == "Good":
                    ws[f'AS{i}'] = "2"
                elif fields[27] == "Average":
                    ws[f'AS{i}'] = "3"
                elif fields[27] == "Poor":
                    ws[f'AS{i}'] = "4"
                elif fields[27] == "Very Poor":
                    ws[f'AS{i}'] = "5"

                # 29	Overall indoor environment
                if fields[28] == "Excellent":
                    ws[f'AT{i}'] = "1"
                elif fields[28] == "Good":
                    ws[f'AT{i}'] = "2"
                elif fields[28] == "Average":
                    ws[f'AT{i}'] = "3"
                elif fields[28] == "Poor":
                    ws[f'AT{i}'] = "4"
                elif fields[28] == "Very Poor":
                    ws[f'AT{i}'] = "5"

                # 30	Stuffy Nose
                # 31	Stuffy nose - Feel better or relief after leaving building?
                col = 29
                xcol1 = "AW"
                xcol2 = "AX"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 32	Dry Throat
                # 33	Dry throat - Feel better or relief after leaving building?
                col = 31
                xcol1 = "AY"
                xcol2 = "AZ"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 34	Cough
                # 35	Cough - Feel better or relief after leaving building?
                col = 33
                xcol1 = "BA"
                xcol2 = "BB"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 36	Skin Rash / Itchiness
                # 37	Skin rash / Itchiness - Feel better or relief after leaving building?
                col = 35
                xcol1 = "BC"
                xcol2 = "BD"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 38	Eye Irritation
                # 39	Eye Irritation - Feel better or relief after leaving building?
                col = 37
                xcol1 = "BE"
                xcol2 = "BF"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 40	Headache
                # 41	Headache - Feel better or relief after leaving building?
                col = 39
                xcol1 = "BG"
                xcol2 = "BH"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 42	Lethargy
                # 43	Lethargy - Feel better or relief after leaving building?
                col = 41
                xcol1 = "BI"
                xcol2 = "BJ"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 44	Drowsiness
                # 45	Drowsiness - Feel better or relief after leaving building?
                col = 43
                xcol1 = "BK"
                xcol2 = "BL"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 46	Dizziness
                # 47	Dizziness - Feel better or relief after leaving building?
                col = 45
                xcol1 = "BM"
                xcol2 = "BN"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 48	Nausea / Vomitting
                # 49	Nausea / Vomitting - Feel better or relief after leaving building?
                col = 47
                xcol1 = "BO"
                xcol2 = "BP"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 50	Shortness of breath
                # 51	Shortness of breath - Feel better or relief after leaving building?
                col = 49
                xcol1 = "BQ"
                xcol2 = "BR"
                ws[f'{xcol1}{i}'] = fields[col]
                ws[f'{xcol2}{i}'] = "NA"
                if fields[col] == "Daily" or fields[col] == "2-3 times weekly":
                    ws[f'{xcol2}{i}'] = fields[col+1]

                # 52	Consume healthier meals. (e.g. choose healthier options when dining out or when catering in)
                col = 51
                xcol = "BU"
                if fields[col] == "> 3 times a week":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "1-3 times a week":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Less than once a week":
                    ws[f'{xcol}{i}'] = "3"
                elif fields[col] == "Never":
                    ws[f'{xcol}{i}'] = "4"

                # 53	Consume sweetened drinks. (e.g. soft drinks, fruit drinks, coffee, tea milo etc)
                col = 52
                xcol = "BV"
                if fields[col] == "> 3 times a week":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "1-3 times a week":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Less than once a week":
                    ws[f'{xcol}{i}'] = "3"
                elif fields[col] == "Never":
                    ws[f'{xcol}{i}'] = "4"

                # 54	Take the stairs instead of the lift.
                col = 53
                xcol = "BW"
                if fields[col] == "> 3 times a week":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "1-3 times a week":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Less than once a week":
                    ws[f'{xcol}{i}'] = "3"
                elif fields[col] == "Never":
                    ws[f'{xcol}{i}'] = "4"

                # 55	Engage in regular physical activity (any form of sports/exercise for at least 20 minutes per session)
                col = 54
                xcol = "BX"
                if fields[col] == "> 3 times a week":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "1-3 times a week":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Less than once a week":
                    ws[f'{xcol}{i}'] = "3"
                elif fields[col] == "Never":
                    ws[f'{xcol}{i}'] = "4"

                # 56	Smoke.
                col = 55
                xcol = "CA"
                if fields[col] == "At least once a day":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "At least once a week":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Never":
                    ws[f'{xcol}{i}'] = "3"

                # 57	Manage your stress levels.
                col = 56
                xcol = "CD"
                if fields[col] == "Very Well":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "Quite Well":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Not Well":
                    ws[f'{xcol}{i}'] = "3"
                elif fields[col] == "Cannot Cope":
                    ws[f'{xcol}{i}'] = "4"

                # 58	How satisfied are you with the health activities organised at your workplace?
                col = 57
                xcol = "CG"
                if fields[col] == "Very Satisfied":
                    ws[f'{xcol}{i}'] = "1"
                elif fields[col] == "Satisfied":
                    ws[f'{xcol}{i}'] = "2"
                elif fields[col] == "Neutral":
                    ws[f'{xcol}{i}'] = "3"
                elif fields[col] == "Not Satisfied":
                    ws[f'{xcol}{i}'] = "4"
                elif fields[col] == "Extremely not satisfied":
                    ws[f'{xcol}{i}'] = "5"
                elif fields[col] == "I am not aware of the health initiatives":
                    ws[f'{xcol}{i}'] = "6"

                # 59	At my work, I feel bursting with energy
                col = 58
                xcol = "CJ"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 60	At my job, I feel strong and vigorous
                col = 59
                xcol = "CK"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 61	I am enthusiastic about my job
                col = 60
                xcol = "CL"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 62	My job inspires me
                col = 61
                xcol = "CM"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 63	When I get up in the morning, I feel like going to work
                col = 62
                xcol = "CN"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 64	I feel happy when I am working intensely.
                col = 63
                xcol = "CO"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 65	I am proud of the work I do
                col = 64
                xcol = "CP"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 66	I am immersed in my work
                col = 65
                xcol = "CQ"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 67	I get carried away when I am working
                col = 66
                xcol = "CR"
                ws[f'{xcol}{i}'] = fields[col].split("-")[0].strip()

                # 68	Do you have any other comments and feedback?
                col = 67
                xcol = "CU"
                if len(fields) > 67:
                    ws[f'{xcol}{i}'] = fields[col]
                else:
                    ws[f'{xcol}{i}'] = ""

                i += 1

    wb.save(output_xlsx_file)









if __name__ == "__main__":
    main() # type: ignore
